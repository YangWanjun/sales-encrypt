import datetime
import re
import openpyxl as px
from openpyxl.writer.excel import save_virtual_workbook

from utils import common, constants
from utils.app_base import get_member_expense_template_path
from utils.errors import CustomException


def gen_attendance_by_month(path_template, results):
    book = px.load_workbook(path_template)
    sheet = book.get_sheet_by_name('Sheet1')

    start_row = 5
    count = len(results)
    set_openpyxl_styles(sheet, start_row, start_row + (count - 1 if count else 0), 2, 21)
    for i, row_data in enumerate(results):
        # NO
        sheet.cell(row=start_row + i, column=2).value = "=ROW() - 4"
        # 隠し項目（Content Type ID)
        sheet.cell(row=start_row + i, column=3).value = row_data.get('content_type_id') or ''
        # 隠し項目（Object ID)
        sheet.cell(row=start_row + i, column=4).value = row_data.get('object_id') or ''
        # 隠し項目（Project ID)
        sheet.cell(row=start_row + i, column=5).value = row_data.get('project_id') or ''
        # 社員番号
        sheet.cell(row=start_row + i, column=6).value = row_data.get('code') or ''
        # 氏名
        sheet.cell(row=start_row + i, column=7).value = row_data.get('full_name') or ''
        # 所在部署
        sheet.cell(row=start_row + i, column=8).value = row_data.get('organization_name') or ''
        # 所属会社
        sheet.cell(row=start_row + i, column=9).value = row_data.get('company_name') or ''
        # 契約形態
        sheet.cell(row=start_row + i, column=10).value = common.get_name_from_choice(
            row_data.get('contract_type'), choice=constants.CHOICE_CONTRACT_TYPE
        ) or ''
        # 案件名
        sheet.cell(row=start_row + i, column=11).value = row_data.get('project_name') or ''
        # 最寄駅
        sheet.cell(row=start_row + i, column=12).value = row_data.get('nearest_station') or ''
        # 顧客会社
        sheet.cell(row=start_row + i, column=13).value = row_data.get('client_name') or ''
        # 契約種類
        sheet.cell(row=start_row + i, column=14).value = "一括" if row_data.get('is_blanket_contract') == 1 else 'SES'
        # 勤務時間
        sheet.cell(row=start_row + i, column=15).value = row_data.get('total_hours') or ''
        # ＢＰ勤務時間
        sheet.cell(row=start_row + i, column=16).value = row_data.get('total_hours_bp') or ''
        # 勤務日数
        sheet.cell(row=start_row + i, column=17).value = row_data.get('total_days') or ''
        # 深夜日数
        sheet.cell(row=start_row + i, column=18).value = row_data.get('night_days') or ''
        # 客先立替金
        sheet.cell(row=start_row + i, column=19).value = ''
        # 立替金
        sheet.cell(row=start_row + i, column=20).value = ''
        # 通勤交通費
        sheet.cell(row=start_row + i, column=21).value = row_data.get('commuting_amount') or ''
    return save_virtual_workbook(book)


def gen_worker_roster(path_template, member):
    contract = member.contracts.filter(end_date__gte=datetime.date.today()).order_by('start_date').first()
    book = px.load_workbook(path_template)
    # 氏名
    set_named_range_value(book, 'POS_NAME', member.full_name)
    # フリガナ
    set_named_range_value(book, 'POS_NAME_FURIGANA', member.full_kana)
    # 生年月日
    set_named_range_value(book, 'POS_BIRTHDAY', member.birthday)
    # 性別
    set_named_range_value(book, 'POS_GENDER', member.get_gender_display())
    if contract:
        # 業務の種類
        set_named_range_value(book, 'POS_BUSINESS_TYPE', contract.get_business_type())
    # 社員住所の郵便番号
    set_named_range_value(book, 'POS_POSTCODE', member.post_code)
    # 社員の携帯番号
    set_named_range_value(book, 'POS_TEL', member.phone)
    # 現住所
    set_named_range_value(book, 'POS_ADDRESS', member.address)
    if member.residence:
        # 在留資格
        if member.residence.residence_type:
            set_named_range_value(book, 'POS_RESIDENCE_STATUS', member.residence.get_residence_type_display())
        if member.residence.expired_date:
            expired_date = member.residence.expired_date.strftime('%Y 年 %m 月 %d 日')
            set_named_range_value(book, 'POS_RESIDENCE_EXPIRED_DATE', expired_date)
        set_named_range_value(book, 'POS_RESIDENCE_NO', member.residence.residence_no)
    if member.join_date:
        # 雇入年月日
        set_named_range_value(book, 'POS_JOIN_DATE', member.join_date.strftime('%Y 年 %m 月 %d 日'))
    # 勤務交通費
    set_named_range_value(book, 'POS_COMMUTATION_AMOUNT', member.get_commutation_amount())
    # 扶養家族情報
    if contract and contract.has_health_insurance:
        qs_family = member.families.filter(is_deleted=False)
    else:
        # 社会保険を加入していない場合、扶養人情報の出力は不要
        qs_family = member.families.none()
    for i, family in enumerate(qs_family, start=1):
        # 最大6件しか入力できない。
        if i > 6:
            break
        # 続柄
        set_named_range_value(book, 'POS_FAMILY_RELATIONSHIP_{}'.format(i), family.get_relationship_display())
        # 氏名
        set_named_range_value(book, 'POS_FAMILY_NAME_SPELL_{}'.format(i), family.kana)
        set_named_range_value(book, 'POS_FAMILY_NAME_{}'.format(i), family.name)
        # 生年月日
        set_named_range_value(book, 'POS_FAMILY_BIRTHDAY_{}'.format(i), family.birthday)
        # 性別
        set_named_range_value(book, 'POS_FAMILY_GENDER_{}'.format(i), family.get_gender_display())
        # 個人番号
        set_named_range_value(book, 'POS_FAMILY_PERSONAL_NUMBER_{}'.format(i), family.personal_number)
        # 同居
        set_named_range_value(book, 'POS_FAMILY_LIVING_TOGETHER_{}'.format(i), '○' if family.is_living_together else '×')
        # 年収
        if family.has_income == '0':
            set_named_range_value(book, 'POS_FAMILY_ANNUAL_INCOME_{}'.format(i), 'なし')
        else:
            set_named_range_value(book, 'POS_FAMILY_ANNUAL_INCOME_{}'.format(i), family.annual_income)
    # 個人番号
    set_named_range_value(book, 'POS_PERSONAL_NUMBER', member.personal_number)
    # 基礎年金番号
    if contract and contract.has_health_insurance:
        basic_pension_no = member.basic_pension_no
    else:
        # 社会保険を加入していない場合、基礎年金番号を入力されても出力しない
        basic_pension_no = None
    set_named_range_value(book, 'POS_BASIC_PENSION_NO', basic_pension_no)
    # 雇用保険証被保険者番号
    set_named_range_value(book, 'POS_EMPLOYMENT_INSURANCE_NO', member.employment_insurance_no)
    return save_virtual_workbook(book)


def gen_member_expense(member, year, month):
    qs_expense = member.expenses.filter(
        is_deleted=False, year=year, month=month, cost_type='01',
    ).order_by('date')
    first_expense = qs_expense.first()
    data_count = qs_expense.count()
    if first_expense is None:
        raise CustomException(constants.ERROR_NO_EXPENSES.format(
            name=str(member),
            year=year,
            month=month,
        ))
    elif data_count > constants.EXPENSE_EXPORT_MAX_ROWS:
        raise CustomException(constants.ERROR_EXPENSE_COUNT_OVER.format(
            name=str(member),
            year=year,
            month=month,
            count=constants.EXPENSE_EXPORT_MAX_ROWS,
        ))
    if data_count <= 40:
        path_template = get_member_expense_template_path()
    else:
        path_template = get_member_expense_template_path(is_row_over=True)
    book = px.load_workbook(path_template)
    # 氏名
    set_named_range_value(book, 'POS_NAME', member.full_name)
    # 作成者
    set_named_range_value(book, 'POS_CREATOR', member.full_name)
    # 所属
    if first_expense.organization:
        organization = first_expense.organization
    else:
        organization = member.get_organization(first_expense.year, first_expense.month)
    set_named_range_value(book, 'POS_ORGANIZATION', organization and organization.name)
    # 申請日
    first_day = datetime.date(int(year), int(month), 1)
    set_named_range_value(book, 'POS_SUBMIT_DATE', common.get_last_day_by_month(first_day))
    for i, expense in enumerate(qs_expense):
        set_named_range_value(book, 'POS_ITEM_DATE', expense.date, offset_row=i)
        set_named_range_value(book, 'POS_ITEM_CONTENT', expense.content, offset_row=i)
        set_named_range_value(book, 'POS_ITEM_SUBJECT', expense.category.name, offset_row=i)
        set_named_range_value(book, 'POS_ITEM_PAY_DST', expense.payment_destination, offset_row=i)
        set_named_range_value(book, 'POS_ITEM_AMOUNT', expense.amount, offset_row=i)
        set_named_range_value(book, 'POS_ITEM_TAX_RATE', expense.tax_rate, offset_row=i)
        set_named_range_value(book, 'POS_ITEM_AMOUNT_NO_TAX', expense.amount_no_tax, offset_row=i)
        set_named_range_value(book, 'POS_ITEM_TAX_AMOUNT', expense.tax_amount, offset_row=i)
        set_named_range_value(book, 'POS_ITEM_COMMENT', expense.comment, offset_row=i)

    return save_virtual_workbook(book)


def set_openpyxl_styles(ws, min_row, max_row, min_col, max_col):
    rows = list(ws.iter_rows(min_row, max_row, min_col, max_col))
    style_list = []
    reg = re.compile(r'\b[A-Z]{1,2}([0-9])+\b')
    dict_formulae = {}

    # we convert iterator to list for simplicity, but it's not memory efficient solution
    rows = list(rows)
    for row_index, cells in enumerate(rows):
        for col_index, cell in enumerate(cells):
            if row_index == 0:
                style_list.append((cell.border.copy(),
                                   cell.font.copy(),
                                   cell.fill.copy(),
                                   cell.alignment.copy(),
                                   cell.number_format))
                # フォーミュラ
                if cell.value and str(cell.value)[0] == '=':
                    lst = reg.findall(cell.value)
                    if lst and lst.count(lst[0]) == len(lst) and int(lst[0]) == min_row:
                        dict_formulae[col_index] = cell.value.replace(lst[0], '{0}')
            else:
                cell.border = style_list[col_index][0]
                cell.font = style_list[col_index][1]
                cell.fill = style_list[col_index][2]
                cell.alignment = style_list[col_index][3]
                cell.number_format = style_list[col_index][4]
                if cell.value and cell.value[0] == '=':
                    pass
                elif col_index in dict_formulae:
                    formulae = dict_formulae[col_index].format(min_row + row_index)
                    cell.value = formulae


def set_named_range_value(book, name, value, offset_row=0, offset_col=0):
    if name in book.defined_names:
        for sheet_title, sheet_coords in book.defined_names[name].destinations:
            sheet = book[sheet_title]
            if offset_row > 0 or offset_col > 0:
                cell = sheet[sheet_coords]
                sheet.cell(row=cell.row + offset_row, column=cell.column + offset_col, value=value)
            else:
                sheet[sheet_coords].value = value
