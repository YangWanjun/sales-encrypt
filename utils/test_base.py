import codecs
import os
import openpyxl as px

from django.conf import settings
from django.contrib.auth.models import User
from django.db import connection

from rest_framework.test import APITestCase


class BaseAPITestCase(APITestCase):

    @classmethod
    def setUpTestData(cls):
        cls.create_ddl()
        # 初期データ導入
        with connection.cursor() as cur:
            cur.execute('DELETE FROM test_sales.auth_permission;')
            cur.execute('DELETE FROM test_sales.django_content_type;')
        init_data_path = os.path.join(settings.BASE_DIR, 'data/test/data/init_data.xlsx')
        if os.path.exists(init_data_path):
            cls.read_data_from_file(init_data_path)

        User.objects.create_user(
            username='admin',
            email='ywjsailor@gmail.com',
            password='admin',
            is_staff=True,
            is_superuser=True,
            first_name='',
            last_name='管理者',
        )
        with connection.cursor() as cur:
            cur.execute('select count(*) from test_sales.django_content_type;')
            print(cur.fetchone())

    @classmethod
    def create_ddl(cls, *args, **kwargs):
        path = os.path.join(settings.BASE_DIR, 'data', 'SQL')
        with connection.cursor() as cur:
            for name in os.listdir(path):
                if not name.lower().endswith('.sql'):
                    continue
                with codecs.open(os.path.join(path, name), 'r', 'utf-8') as f:
                    sql_list = f.read().split('//')
                if len(sql_list) == 1:
                    cur.execute(sql_list[0])
                else:
                    for i, sql in enumerate(sql_list):
                        if sql.strip().upper().startswith('DELIMITER'):
                            continue
                        else:
                            cur.execute(sql)

    @classmethod
    def read_data_from_file(cls, path):
        book = px.load_workbook(path, read_only=True)
        for sheet in book.worksheets:
            table_name = sheet.cell(2, 2).value
            columns = []
            data_types = []
            values_list = []
            for i, row in enumerate(sheet.iter_rows(), start=1):
                if i < 12:
                    continue
                elif i == 12:
                    columns = [cell.value for cell in row if cell.value]
                elif i == 13:
                    data_types = [cell.value for cell in row if cell.value]
                else:
                    values_list.append([None if cell.value == '« NULL »' else cell.value for cell in row if cell.value is not None])
            with connection.cursor() as cur:
                for values in values_list:
                    sql = 'INSERT INTO test_sales.{} ({}) values ({});'.format(
                        table_name,
                        ','.join(columns),
                        ','.join(['%s'] * len(columns))
                    )
                    try:
                        cur.execute(sql, values)
                    except Exception as ex:
                        print('↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓')
                        print(sql)
                        print(values)
                        print('↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑')
                        raise ex
